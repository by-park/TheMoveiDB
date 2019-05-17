# pip install openpyxl 필요
# 참고 사이트: http://pythonstudy.xyz/python/article/405-%ED%8C%8C%EC%9D%B4%EC%8D%AC-%EC%97%91%EC%85%80-%EC%82%AC%EC%9A%A9%ED%95%98%EA%B8%B0
# id, title_ko, title_en, year, genres 데이터 들어있는 엑셀 파일

import openpyxl
 
# 엑셀파일 열기
wb = openpyxl.load_workbook('C:/Users/student/Downloads/Moveidbexcel_190514_rating.xlsx')
 
# 현재 Active Sheet 얻기
ws = wb.active
# ws = wb.get_sheet_by_name("Sheet1")

# 장르 목록
genres_list = ['액션', '어드벤처', '애니메이션', '코미디', '범죄', '다큐멘터리', '드라마',
               '가족', '판타지', '호러', '뮤지컬', '미스터리', '로맨스', 'SF', '스릴러', '전쟁', '서부']

# 영화 정보를 읽기
for r in ws.rows:
    row_index = r[0].row   # 행 인덱스
    title_ko = r[1].value
    genres = r[4].value
 
    # 장르 나누기
    for g_id in range(len(genres_list)):
        if genres_list[g_id] in genres:
            ws.cell(row=row_index, column=5+g_id).value = 1
        else:
            ws.cell(row=row_index, column=5+g_id).value = 0
 
    print(row_index, title_ko, genres)
 
# 엑셀 파일 저장
wb.save("C:/Users/student/Downloads/Moveidbexcel_divided_genres.xlsx")
wb.close()
